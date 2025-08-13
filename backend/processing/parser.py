# backend/processing/parser.py (updated for new Excel format)

import pandas as pd
from io import BytesIO
import datetime
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import os

def format_date(date_str):
    try:
        # Parse the date string
        date_obj = pd.to_datetime(date_str)
        # Format the date with ordinal suffix
        day = date_obj.day
        suffix = 'th' if 11 <= day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
        return f"{day}{suffix} {date_obj.strftime('%B %Y')}"
    except:
        return date_str

def format_time(time_str):
    try:
        # Parse the time string
        time_obj = datetime.datetime.strptime(time_str, '%H:%M:%S')
        # Format in 12-hour format with AM/PM
        return time_obj.strftime('%I:%M %p')
    except:
        return time_str

def parse_excel(file):
    xls = pd.ExcelFile(file)

    # ---- Extract Metadata from 'Info' ----
    info_sheet = xls.parse("Info", header=None)
    metadata = {
        "year": "",
        "semester": "",
        "exam_type": "",
        "department": "",
        "program_type": "",
        "subject_code": "",
        "subject_name": "",
        "lecturer": "",
        "date": "",
        "time": "",
        "exam_type_code": ""
    }

    for i in range(len(info_sheet)):
        row = info_sheet.iloc[i]
        for cell in row:
            if isinstance(cell, str):
                cell_l = cell.lower()
                if "year" in cell_l:
                    metadata["year"] = str(row[row[row == cell].index[0] + 1])
                elif "semester" in cell_l:
                    metadata["semester"] = str(row[row[row == cell].index[0] + 1])
                elif "exam type" in cell_l:
                    metadata["exam_type"] = str(row[row[row == cell].index[0] + 1])
                elif "department" in cell_l:
                    metadata["department"] = str(row[row[row == cell].index[0] + 1])
                elif "program type" in cell_l:
                    metadata["program_type"] = str(row[row[row == cell].index[0] + 1])
                elif "subject code" in cell_l:
                    metadata["subject_code"] = str(row[row[row == cell].index[0] + 1])
                elif "subject name" in cell_l:
                    metadata["subject_name"] = str(row[row[row == cell].index[0] + 1])
                elif "lecturer" in cell_l:
                    metadata["lecturer"] = str(row[row[row == cell].index[0] + 1])
                elif "date" in cell_l:
                    metadata["date"] = format_date(str(row[row[row == cell].index[0] + 1]))
                elif "time" in cell_l:
                    time_start = format_time(str(row[row[row == cell].index[0] + 1]))
                    time_end = format_time(str(row[row[row == cell].index[0] + 2]))
                    metadata["time"] = f"{time_start} - {time_end}"

    # Construct exam_type_code after all fields are populated
    metadata["exam_type_code"] = f"{metadata['exam_type']}_{metadata['semester']}/{metadata['year']}"

    # ---- Extract Questions ----
    all_questions = []

    # Multiple Choice
    df_mc = xls.parse("MultipleChoice")
    # Load workbook and sheet with openpyxl for image extraction
    wb = openpyxl.load_workbook(file)
    ws = wb["MultipleChoice"]
    image_loader = SheetImageLoader(ws)
    # Map column names to indices
    col_map = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 0)}
    images_dir = os.path.join(os.path.dirname(__file__), "templates", "images")
    os.makedirs(images_dir, exist_ok=True)
    # Print all coordinates where images are found
    print(f"[DEBUG] ImageLoader found images at: {list(image_loader._images.keys())}")
    for i, row in enumerate(df_mc.itertuples(index=False), 2):  # DataFrame is 0-based, Excel is 1-based, skip header
        # Check if any option has length >= 20
        options = [str(getattr(row, opt, "")) for opt in ['a', 'b', 'c', 'd', 'e']]
        is_long = any(len(opt) >= 20 for opt in options)
        # Extract image if present
        image_path = ""
        if "image" in col_map:
            excel_col_idx = col_map["image"]
            cell = ws.cell(row=i, column=excel_col_idx+1)  # openpyxl is 1-based
            print(f"[DEBUG] Checking cell for image: {cell.coordinate}")
            if image_loader.image_in(cell.coordinate):
                img = image_loader.get(cell.coordinate)
                image_path = os.path.join(images_dir, f"mcq_{i}.png")
                img.save(image_path)
                image_path = os.path.relpath(image_path, os.path.dirname(__file__))
                print(f"[DEBUG] Parsed image for MCQ row {i}: {image_path}")
            else:
                val = cell.value
                if val and isinstance(val, str):
                    image_path = val
                    print(f"[DEBUG] Parsed legacy image path for MCQ row {i}: {image_path}")
        all_questions.append({
            "type": "multiple choice",
            "question": getattr(row, "question", ""),
            "a": getattr(row, "a", ""),
            "b": getattr(row, "b", ""),
            "c": getattr(row, "c", ""),
            "d": getattr(row, "d", ""),
            "e": getattr(row, "e", ""),
            "answer": getattr(row, "ans", ""),
            "category": getattr(row, "category", ""),
            "image": image_path,
            "is_long": is_long
        })

    # True/False
    df_tf = xls.parse("TrueFalse")
    for _, row in df_tf.iterrows():
        all_questions.append({
            "type": "true/false",
            "question": row.get("question", ""),
            "answer": row.get("ans", ""),
            "category": row.get("category", "")
        })

    # Matching
    df_match = xls.parse("Matching")
    for _, row in df_match.iterrows():
       all_questions.append({
            "type": "matching",
            "question": row.get("question", ""),
            "answer": row.get("ans", ""),
            "category": row.get("category", "")
        })

    # Fake Answers (for matching questions distractors)
    try:
        df_fake = xls.parse("FakeAnswers")
        for _, row in df_fake.iterrows():
            all_questions.append({
                "type": "fake answer",
                "question": row.get("question", ""),
                "answer": row.get("ans", ""),
                "category": row.get("category", "")
            })
    except Exception:
        # If FakeAnswers sheet doesn't exist, skip it
        pass

    # Written Question
    df_written = xls.parse("WrittenQuestion")
    for _, row in df_written.iterrows():
        all_questions.append({
            "type": "written question",
            "question": row.get("question", ""),
            "answer": row.get("ans", ""),
            "q_type": row.get("q_type", ""),
            "category": row.get("category", "")
        })

    metadata["selection_settings"] = {}  # Optional: add default/random if needed
    return all_questions, metadata
